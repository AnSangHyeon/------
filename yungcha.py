# 필요한 selenium,openpyxl,os,bs4라이브러리를 가져온다
from selenium import webdriver
from openpyxl import Workbook
from openpyxl.styles import Alignment
import os
from bs4 import BeautifulSoup
import openpyxl

# Selenium을 사용하여 브라우저를 제어하기 위한 웹 드라이버를 설정
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(options = option)

# 기본경로를 설정해준다 
url = 'https://upbit.com/exchange?code=CRIX.UPBIT.KRW-BTC'

# url(변수)에 있는 path로 연결
driver.get(url)

# openpyxl을 사용하기위한 설정을 한다
wb = openpyxl.Workbook()
current_directory = os.getcwd()
file_path = os.path.join(current_directory, 'yungcha.xlsx')
ws = wb.active 

# soup(변수)에 문서(HTML,XML)를 가져와서 할당해준다
soup = BeautifulSoup(driver.page_source, 'html.parser')

# 크롤링해온 값들을 담아줄 배열을 선언한다
namearr = []
pricearr = []
comparedToThePreviousDay1arr = []
comparedToThePreviousDay2arr = []
transactionAmountarr = []

# 해더에 넣을 키워드들을 선언하여 중앙정렬을 선언하고 for문으로 키워드들을 중앙정렬하여 넣어준다
headerValues = ['한글명', '현재가', '전일대비', '전일대비', '거래대금']
for index, headerText in enumerate(headerValues, start=1):
  ws.cell(row=1, column=index, value=headerText).alignment = Alignment(horizontal='center')

#datas변수에 가져올 data들을 묶어놓은 tr(element)를 가져온다
datas = soup.select('#UpbitLayout > div:nth-child(4) > div > section.ty02 > article > span.tabB > div > div > div:nth-child(1) > table > tbody > tr')

# 가져온 data로 반복문을 돌려서 
# strongTag(변수)에는 한글명 
# price(변수)에는 현재가 
# comparedToThePreviousDay1(변수)에는 전일대비중 위에값 
# comparedToThePreviousDay2(변수)에는 전일대비중 아래값 
# transactionAmount(변수)에는 거래대금을 넣어주고 
# 위에 만들어뒀던 배열에 append한다
for stock in datas:
  strongTag = stock.select_one('td.tit > a > strong')
  if strongTag is not None:
    price = stock.select_one('td.price > strong').text
    name = strongTag.text
    comparedToThePreviousDay1 = stock.select_one('td.percent > p').text
    comparedToThePreviousDay2 = stock.select_one('td.percent > em').text
    transactionAmount = stock.select_one('td.rAlign > p').text
    namearr.append(name)
    pricearr.append(price)
    comparedToThePreviousDay1arr.append(comparedToThePreviousDay1)
    comparedToThePreviousDay2arr.append(comparedToThePreviousDay2)
    transactionAmountarr.append(transactionAmount)
 
# 엑셀에 넣었을때 글자가 안보여서 넓이를 넓혀준다
ws.column_dimensions['A'].width = 14
ws.column_dimensions['B'].width = 11
ws.column_dimensions['C'].width = 11
ws.column_dimensions['E'].width = 11

# 배열에 넣은값들을 엑셀에 정렬하여 넣어준다 
for index, value in enumerate(namearr, start=2):
  ws.cell(row=index, column=1, value=value).alignment = Alignment(horizontal='center')
for index, value in enumerate(pricearr, start=2):
  ws.cell(row=index, column=2, value=value).alignment = Alignment(horizontal='right')
for index, value in enumerate(comparedToThePreviousDay1arr, start=2):
  ws.cell(row=index, column=3, value=value).alignment = Alignment(horizontal='right')
for index, value in enumerate(comparedToThePreviousDay2arr, start=2):
  ws.cell(row=index, column=4, value=value).alignment = Alignment(horizontal='right')
for index, value in enumerate(transactionAmountarr, start=2):
  ws.cell(row=index, column=5, value=value).alignment = Alignment(horizontal='right')
  
# 엑셀파일에 저장해준다  
wb.save(file_path) 

# 크롤링이 끝났으니 실행을 종료한다
driver.quit()